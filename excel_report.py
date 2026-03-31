"""
excel_report.py
Builds a daily Excel report of all cellular plans.
One sheet per carrier; rows changed in the last 24h are highlighted yellow.
"""
import io
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import get_plans, get_changes

CARRIER_NAMES = {
    "partner":   "פרטנר",
    "pelephone": "פלאפון",
    "hotmobile": "הוט מובייל",
    "cellcom":   "סלקום",
    "mobile019": "019",
}

CARRIER_ORDER = ["partner", "pelephone", "hotmobile", "cellcom", "mobile019"]

CHANGE_LABELS = {
    "price_change":  "שינוי מחיר",
    "new_plan":      "חבילה חדשה",
    "removed_plan":  "הוסרה",
    "extras_change": "שינוי הטבות",
}

HEADERS = ["שם חבילה", "מחיר ₪", "גיגה", "דקות", "הטבות", "שינוי ב-24ש׳", "ערך ישן"]
COL_WIDTHS = [28, 10, 8, 14, 38, 22, 14]

# Styles
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CHANGED_FILL = PatternFill("solid", fgColor="FFFF00")
CHANGED_FONT = Font(bold=True, color="000000", size=11)
NORMAL_FONT  = Font(color="000000", size=11)
THIN_BORDER  = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT  = Alignment(horizontal="right",  vertical="center", wrap_text=True)


def _recent_changes_lookup() -> dict:
    """Return {(carrier, plan_name): change_dict} for changes in last 24h."""
    cutoff = datetime.now() - timedelta(hours=24)
    all_changes = get_changes(limit=1000)
    lookup = {}
    for ch in all_changes:
        try:
            changed_at = datetime.fromisoformat(ch["changed_at"])
        except (ValueError, TypeError):
            continue
        if changed_at >= cutoff:
            key = (ch["carrier"], ch["plan_name"])
            # Keep first (most recent) entry per plan
            if key not in lookup:
                lookup[key] = ch
    return lookup


def _write_sheet(ws, plans: list, changes_lookup: dict) -> None:
    ws.sheet_view.rightToLeft = True
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 22

    # Data rows
    for row_idx, plan in enumerate(plans, start=2):
        key = (plan["carrier"], plan["plan_name"])
        change = changes_lookup.get(key)
        is_changed = change is not None

        gb_val = plan.get("data_gb")
        gb_display = "ללא הגבלה" if gb_val is None else gb_val
        extras_str = ", ".join(plan.get("extras") or [])
        change_label = CHANGE_LABELS.get(change["change_type"], "") if change else ""
        old_val = change.get("old_val", "") if change else ""

        row_data = [
            plan.get("plan_name", ""),
            plan.get("price", ""),
            gb_display,
            plan.get("minutes", ""),
            extras_str,
            change_label,
            old_val,
        ]

        fill = CHANGED_FILL if is_changed else PatternFill()
        font = CHANGED_FONT if is_changed else NORMAL_FONT

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = fill
            cell.font      = font
            cell.alignment = RIGHT
            cell.border    = THIN_BORDER
        ws.row_dimensions[row_idx].height = 18


def build_excel_report(db_path=None) -> bytes:
    """Build Excel workbook and return as bytes ready for email attachment."""
    all_plans   = get_plans(db_path=db_path)
    changes_lkp = _recent_changes_lookup()

    # Group plans by carrier
    by_carrier: dict = {k: [] for k in CARRIER_ORDER}
    for plan in all_plans:
        carrier = plan["carrier"]
        if carrier in by_carrier:
            by_carrier[carrier].append(plan)

    wb = Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for carrier_key in CARRIER_ORDER:
        sheet_name = CARRIER_NAMES.get(carrier_key, carrier_key)
        ws = wb.create_sheet(title=sheet_name)
        _write_sheet(ws, by_carrier.get(carrier_key, []), changes_lkp)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
